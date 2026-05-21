#!/usr/bin/env python3
"""
Overleaf Comment Scraper
Parses a saved Overleaf page HTML file and exports review comments to a spreadsheet.

Usage:
    python3 scraper.py <input.html> [output.xlsx]

The input HTML should be produced by:
    - Firefox: right-click the review panel area → "View Selection Source" (captures rendered DOM)
    - OR: Save a fully rendered page snapshot that visibly contains the review panel entries

Outputs an Excel workbook (.xlsx) and optionally a CSV with columns:
  Thread ID | Author | Date | Comment | Highlighted Text | Context
"""

import sys
import re
import csv
from pathlib import Path
from bs4 import BeautifulSoup, NavigableString

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False


# ---------------------------------------------------------------------------
# DOM extraction helpers
# ---------------------------------------------------------------------------

def extract_user(entry_el):
    """Return the display name from a review-panel-entry element."""
    user_div = entry_el.find(class_="review-panel-entry-user")
    if not user_div:
        return ""
    # The div contains a badge <span> followed by the text node with the name.
    texts = [t for t in user_div.children if isinstance(t, NavigableString)]
    name = "".join(texts).strip()
    if not name:
        # Fallback: full text minus badge aria content
        name = user_div.get_text(separator="", strip=True)
    return name


def extract_thread_id(entry_el):
    """Extract the thread/comment ID from the options button id attribute."""
    btn = entry_el.find(id=re.compile(r"review-panel-comment-options-btn-(.+)"))
    if btn:
        return btn["id"].replace("review-panel-comment-options-btn-", "")
    return ""


def extract_comments_from_entry(entry_el):
    """
    Return a list of (author, date, body) tuples for all comments in a thread
    entry, including replies.
    """
    results = []
    # Each individual comment inside a thread is in .review-panel-comment
    for comment_el in entry_el.find_all(class_="review-panel-comment"):
        author = extract_user(comment_el)
        time_el = comment_el.find(class_="review-panel-entry-time")
        date = time_el.get_text(strip=True) if time_el else ""
        body_el = comment_el.find(class_="review-panel-comment-body")
        body = body_el.get_text(separator=" ", strip=True) if body_el else ""
        if body:
            results.append((author, date, body))
    return results


# ---------------------------------------------------------------------------
# Editor text reconstruction
# ---------------------------------------------------------------------------

def build_editor_text(soup):
    """
    Reconstruct the plain-text content of the CodeMirror editor by concatenating
    all .cm-line div text nodes (preserving newlines between lines).
    Returns a list of (line_start_char_offset, line_text) tuples.
    """
    cm_content = soup.find(class_="cm-content")
    if not cm_content:
        return []

    lines = []
    offset = 0
    for line_div in cm_content.find_all(class_="cm-line", recursive=True):
        # Skip nested duplicates (e.g. indent-marker wrappers that are themselves cm-line)
        # by only taking direct children lines of the cm-content
        if line_div.parent != cm_content and line_div.find_parent(class_="cm-line"):
            continue
        text = line_div.get_text(separator="")
        lines.append((offset, text))
        offset += len(text) + 1  # +1 for the implicit newline between lines
    return lines


def get_context_at_pos(lines, pos, window=200):
    """
    Given a list of (start_offset, line_text) tuples and a character position,
    return (highlighted_line, context_snippet).

    highlighted_line: the full line that contains `pos`
    context_snippet:  up to `window` chars of text centred on `pos`
    """
    if not lines:
        return "", ""

    # Find which line contains `pos`
    highlighted_line = ""
    for i, (start, text) in enumerate(lines):
        end = start + len(text)
        if start <= pos <= end:
            highlighted_line = text
            break
    else:
        # pos is beyond end of reconstructed text – use last line
        _, highlighted_line = lines[-1]

    # Rebuild full text for context window
    full_text = "\n".join(t for _, t in lines)
    half = window // 2
    ctx_start = max(0, pos - half)
    ctx_end = min(len(full_text), pos + half)
    context = full_text[ctx_start:ctx_end].strip()

    return highlighted_line, context


# ---------------------------------------------------------------------------
# Main parsing
# ---------------------------------------------------------------------------

def parse_html(html_path):
    """
    Parse an Overleaf HTML file and return a list of row dicts, one per
    individual comment (replies are flattened as separate rows sharing the
    same thread ID).
    """
    html_path = Path(html_path)
    with open(html_path, encoding="utf-8") as f:
        soup = BeautifulSoup(f, "html.parser")

    editor_lines = build_editor_text(soup)

    rows = []

    # Each top-level comment thread entry
    thread_entries = soup.find_all(
        lambda tag: tag.has_attr("class")
        and "review-panel-entry-comment" in tag.get("class", [])
        and tag.name in ("div", "section", "article")
    )

    for entry in thread_entries:
        thread_id = extract_thread_id(entry)
        try:
            data_pos = int(entry.get("data-pos", -1))
        except (ValueError, TypeError):
            data_pos = -1

        highlighted_line, context = get_context_at_pos(editor_lines, data_pos)
        comments = extract_comments_from_entry(entry)

        for idx, (author, date, body) in enumerate(comments):
            rows.append({
                "thread_id": thread_id,
                "comment_index": idx,          # 0 = root, 1+ = replies
                "author": author,
                "date": date,
                "comment": body,
                "highlighted_text": highlighted_line if idx == 0 else "",
                "context": context if idx == 0 else "",
                "char_pos": data_pos if idx == 0 else "",
            })

    return rows


def detect_shell_only_snapshot(html_path):
    """
    Detect an Overleaf shell page where client-side JS was not hydrated into DOM.

    In this case, the file contains loading scaffold and script tags, but no
    editor/comment nodes (no .cm-content, no .review-panel-entry-comment).
    """
    text = Path(html_path).read_text(encoding="utf-8", errors="ignore")
    has_overleaf_shell = (
        "id=\"ide-root\"" in text
        and "loading-screen" in text
        and "pages/ide" in text
    )
    has_rendered_editor = "cm-content" in text
    has_rendered_comments = "review-panel-entry-comment" in text
    return has_overleaf_shell and not (has_rendered_editor or has_rendered_comments)


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

COLUMNS = [
    ("thread_id",       "Thread ID"),
    ("comment_index",   "Reply #"),
    ("author",          "Author"),
    ("date",            "Date"),
    ("comment",         "Comment"),
    ("highlighted_text","Highlighted Text"),
    ("context",         "Context"),
    ("char_pos",        "Char Position"),
]


def write_csv(rows, out_path):
    out_path = Path(out_path)
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[k for k, _ in COLUMNS])
        writer.writerow({k: label for k, label in COLUMNS})
        writer.writerows(rows)
    print(f"CSV written: {out_path}")


def write_excel(rows, out_path):
    if not HAVE_OPENPYXL:
        print("openpyxl not installed – skipping Excel output. Run: pip install openpyxl")
        return

    out_path = Path(out_path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comments"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    # Header row
    for col_idx, (key, label) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))
            cell.alignment = wrap
        # Shade reply rows slightly differently
        if row.get("comment_index", 0) > 0:
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    "solid", fgColor="DCE6F1"
                )

    # Column widths (heuristic)
    col_widths = {
        "Thread ID": 28, "Reply #": 8, "Author": 18, "Date": 18,
        "Comment": 45, "Highlighted Text": 45, "Context": 55, "Char Position": 14,
    }
    for col_idx, (_, label) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = col_widths.get(label, 20)

    ws.freeze_panes = "A2"
    wb.save(out_path)
    print(f"Excel written: {out_path}")


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    html_file = sys.argv[1]
    stem = Path(html_file).stem

    out_excel = sys.argv[2] if len(sys.argv) > 2 else f"{stem}_comments.xlsx"
    out_csv = Path(out_excel).with_suffix(".csv")

    rows = parse_html(html_file)

    if not rows:
        if detect_shell_only_snapshot(html_file):
            print(
                "No comments found: this file looks like an Overleaf shell page (JS not rendered in saved DOM).\n"
                "Use 'View Selection Source' on the open Review panel, or save after the review panel entries are present in the HTML snapshot."
            )
        else:
            print("No comments found. Make sure the Review panel was open and visible when you saved the page.")
        sys.exit(0)

    print(f"Found {len(rows)} comment(s) across {len({r['thread_id'] for r in rows})} thread(s).")

    write_csv(rows, out_csv)
    write_excel(rows, out_excel)


if __name__ == "__main__":
    main()
